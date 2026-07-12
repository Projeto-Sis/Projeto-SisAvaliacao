from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def find_header_row(ws):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        labels = [
            str(ws.cell(row_idx, col_idx).value or "").strip().lower()
            for col_idx in range(1, ws.max_column + 1)
        ]
        if "id" in labels and "status" in labels and "fonte" in labels:
            return row_idx
    raise RuntimeError(f"Nao encontrei cabecalho valido na aba {ws.title!r}.")


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Uso: export_amostras_csv.py entrada.xlsx saida.csv")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    wb = load_workbook(input_path, data_only=True)
    ws = wb["Base_Refinada"] if "Base_Refinada" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row = find_header_row(ws)
    headers = [clean_value(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]

    # Remove colunas finais vazias, sem derrubar colunas intermediarias.
    last_col = len(headers)
    while last_col > 0 and headers[last_col - 1] == "":
        last_col -= 1
    headers = headers[:last_col]

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = [clean_value(ws.cell(row_idx, col).value) for col in range(1, last_col + 1)]
        if any(value != "" for value in row):
            rows.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL)
        writer.writerow(headers)
        writer.writerows(rows)

    print(f"aba={ws.title}")
    print(f"cabecalho_linha={header_row}")
    print(f"colunas={len(headers)}")
    print(f"linhas_dados={len(rows)}")
    print(f"saida={output_path}")


if __name__ == "__main__":
    main()
