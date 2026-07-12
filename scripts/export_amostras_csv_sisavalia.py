from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    ("fonte", "Fonte"),
    ("preco", "Preco (R$)"),
    ("area", "Area (m2)"),
    ("local", "Local (1-3)"),
    ("padrao", "Padrao (1-3)"),
    ("conservacao", "Conservacao (1-3)"),
    ("status_validacao", "status_validacao"),
    ("observacao_validacao", "observacao_validacao"),
]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def find_header_row(ws):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        labels = [
            str(ws.cell(row_idx, col_idx).value or "").strip()
            for col_idx in range(1, ws.max_column + 1)
        ]
        lowered = {label.lower(): idx for idx, label in enumerate(labels)}
        if "fonte" in lowered and "preco (r$)" in lowered and "area (m2)" in lowered:
            return row_idx, labels
    raise RuntimeError(f"Nao encontrei cabecalho valido na aba {ws.title!r}.")


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Uso: export_amostras_csv_sisavalia.py entrada.xlsx saida.csv")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    wb = load_workbook(input_path, data_only=True)
    ws = wb["Base_Refinada"] if "Base_Refinada" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, labels = find_header_row(ws)
    indexes = {label.lower(): idx + 1 for idx, label in enumerate(labels)}

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = []
        for _, original_header in EXPECTED_COLUMNS:
            if original_header == "Fonte":
                source = clean(ws.cell(row_idx, indexes["fonte"]).value)
                reference = clean(ws.cell(row_idx, indexes["local / referencia"]).value)
                url = clean(ws.cell(row_idx, indexes["url fonte"]).value)
                value = " | ".join(str(part) for part in [source, reference, url] if part != "")
            else:
                value = clean(ws.cell(row_idx, indexes[original_header.lower()]).value)
            row.append(value)
        if any(value != "" for value in row):
            rows.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL)
        writer.writerow([column for column, _ in EXPECTED_COLUMNS])
        writer.writerows(rows)

    print(f"aba={ws.title}")
    print(f"linhas_dados={len(rows)}")
    print(f"saida={output_path}")


if __name__ == "__main__":
    main()
